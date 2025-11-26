# src/python_mastery_hub/utils/data_exporters.py
"""
Data Export Utilities - Export Learning Data in Various Formats

Provides utilities for exporting user progress, statistics, and achievements
in different formats including CSV, JSON, PDF reports, and Excel files.
"""

import base64
import csv
import io
import json
import logging
import tempfile
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional, Union

logger = logging.getLogger(__name__)


@dataclass
class ExportConfig:
    """Configuration for data export."""

    include_personal_info: bool = True
    include_progress_data: bool = True
    include_achievements: bool = True
    include_statistics: bool = True
    include_learning_history: bool = True
    date_format: str = "%Y-%m-%d %H:%M:%S"
    timezone: Optional[str] = None


class DataExporter:
    """Base class for data exporters."""

    def __init__(self, config: Optional[ExportConfig] = None):
        self.config = config or ExportConfig()

    def export(
        self, data: Dict[str, Any], output_path: Optional[Path] = None
    ) -> Union[str, bytes]:
        """
        Export data in the specific format.

        Args:
            data: Data to export
            output_path: Optional file path to save

        Returns:
            Exported data as string or bytes
        """
        raise NotImplementedError

    def _format_datetime(self, dt: Union[datetime, str]) -> str:
        """Format datetime consistently."""
        if isinstance(dt, str):
            try:
                dt = datetime.fromisoformat(dt)
            except ValueError:
                return dt

        if isinstance(dt, datetime):
            return dt.strftime(self.config.date_format)

        return str(dt)

    def _filter_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Filter data based on export configuration."""
        filtered = {}

        if self.config.include_personal_info and "user_info" in data:
            filtered["user_info"] = data["user_info"]

        if self.config.include_progress_data and "progress" in data:
            filtered["progress"] = data["progress"]

        if self.config.include_achievements and "achievements" in data:
            filtered["achievements"] = data["achievements"]

        if self.config.include_statistics and "statistics" in data:
            filtered["statistics"] = data["statistics"]

        if self.config.include_learning_history and "learning_history" in data:
            filtered["learning_history"] = data["learning_history"]

        return filtered


class CSVExporter(DataExporter):
    """Export data to CSV format."""

    def export(self, data: Dict[str, Any], output_path: Optional[Path] = None) -> str:
        """Export data to CSV format."""
        filtered_data = self._filter_data(data)
        output = io.StringIO()

        # Export each data section to separate CSV tables
        csv_content = []

        # User information
        if "user_info" in filtered_data:
            csv_content.append(self._export_user_info_csv(filtered_data["user_info"]))

        # Progress data
        if "progress" in filtered_data:
            csv_content.append(self._export_progress_csv(filtered_data["progress"]))

        # Achievements
        if "achievements" in filtered_data:
            csv_content.append(
                self._export_achievements_csv(filtered_data["achievements"])
            )

        # Statistics
        if "statistics" in filtered_data:
            csv_content.append(self._export_statistics_csv(filtered_data["statistics"]))

        # Learning history
        if "learning_history" in filtered_data:
            csv_content.append(
                self._export_learning_history_csv(filtered_data["learning_history"])
            )

        result = "\n\n".join(csv_content)

        if output_path:
            output_path.write_text(result, encoding="utf-8")

        return result

    def _export_user_info_csv(self, user_info: Dict[str, Any]) -> str:
        """Export user information as CSV."""
        output = io.StringIO()
        output.write("# User Information\n")

        writer = csv.writer(output)
        writer.writerow(["Field", "Value"])

        for key, value in user_info.items():
            if isinstance(value, datetime):
                value = self._format_datetime(value)
            writer.writerow([key, str(value)])

        return output.getvalue()

    def _export_progress_csv(self, progress: Dict[str, Any]) -> str:
        """Export progress data as CSV."""
        output = io.StringIO()
        output.write("# Progress Data\n")

        writer = csv.writer(output)

        # Module progress
        if "modules" in progress:
            writer.writerow(
                [
                    "Module",
                    "Completed Topics",
                    "Total Topics",
                    "Percentage",
                    "Time Spent (minutes)",
                ]
            )
            for module_id, module_data in progress["modules"].items():
                writer.writerow(
                    [
                        module_id,
                        module_data.get("completed", 0),
                        module_data.get("total", 0),
                        f"{module_data.get('percentage', 0):.1f}%",
                        module_data.get("time_spent", 0),
                    ]
                )
            output.write("\n")

        # Topic progress
        if "topics" in progress:
            writer.writerow(
                [
                    "Module",
                    "Topic",
                    "Completed",
                    "Completion Date",
                    "Time Spent",
                    "Score",
                ]
            )
            for topic in progress["topics"]:
                writer.writerow(
                    [
                        topic.get("module_id", ""),
                        topic.get("topic_name", ""),
                        "Yes" if topic.get("completed", False) else "No",
                        self._format_datetime(topic.get("completion_date", ""))
                        if topic.get("completion_date")
                        else "",
                        topic.get("time_spent", 0),
                        f"{topic.get('score', 0)*100:.1f}%"
                        if topic.get("score")
                        else "",
                    ]
                )

        return output.getvalue()

    def _export_achievements_csv(self, achievements: List[Dict[str, Any]]) -> str:
        """Export achievements as CSV."""
        output = io.StringIO()
        output.write("# Achievements\n")

        writer = csv.writer(output)
        writer.writerow(
            ["Name", "Description", "Category", "Tier", "Points", "Earned Date"]
        )

        for achievement in achievements:
            writer.writerow(
                [
                    achievement.get("name", ""),
                    achievement.get("description", ""),
                    achievement.get("category", ""),
                    achievement.get("tier", ""),
                    achievement.get("points", 0),
                    self._format_datetime(achievement.get("earned_date", "")),
                ]
            )

        return output.getvalue()

    def _export_statistics_csv(self, statistics: Dict[str, Any]) -> str:
        """Export statistics as CSV."""
        output = io.StringIO()
        output.write("# Learning Statistics\n")

        writer = csv.writer(output)
        writer.writerow(["Metric", "Value"])

        for key, value in statistics.items():
            if isinstance(value, dict):
                # Handle nested dictionaries
                for sub_key, sub_value in value.items():
                    writer.writerow([f"{key}.{sub_key}", str(sub_value)])
            elif isinstance(value, list):
                writer.writerow([key, ", ".join(map(str, value))])
            else:
                writer.writerow([key, str(value)])

        return output.getvalue()

    def _export_learning_history_csv(
        self, learning_history: List[Dict[str, Any]]
    ) -> str:
        """Export learning history as CSV."""
        output = io.StringIO()
        output.write("# Learning History\n")

        writer = csv.writer(output)
        writer.writerow(
            ["Date", "Module", "Topic", "Action", "Duration (minutes)", "Score"]
        )

        for entry in learning_history:
            writer.writerow(
                [
                    self._format_datetime(entry.get("date", "")),
                    entry.get("module_id", ""),
                    entry.get("topic_name", ""),
                    entry.get("action", ""),
                    entry.get("duration", 0),
                    f"{entry.get('score', 0)*100:.1f}%" if entry.get("score") else "",
                ]
            )

        return output.getvalue()


class JSONExporter(DataExporter):
    """Export data to JSON format."""

    def export(self, data: Dict[str, Any], output_path: Optional[Path] = None) -> str:
        """Export data to JSON format."""
        filtered_data = self._filter_data(data)

        # Add export metadata
        export_data = {
            "export_info": {
                "timestamp": datetime.now().isoformat(),
                "format": "json",
                "version": "1.0",
                "exported_by": "Python Mastery Hub",
            },
            "data": filtered_data,
        }

        # Convert datetime objects to strings
        export_data = self._convert_datetimes(export_data)

        result = json.dumps(export_data, indent=2, ensure_ascii=False)

        if output_path:
            output_path.write_text(result, encoding="utf-8")

        return result

    def _convert_datetimes(self, obj: Any) -> Any:
        """Recursively convert datetime objects to strings."""
        if isinstance(obj, datetime):
            return self._format_datetime(obj)
        elif isinstance(obj, dict):
            return {key: self._convert_datetimes(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._convert_datetimes(item) for item in obj]
        else:
            return obj


class HTMLReportExporter(DataExporter):
    """Export data as HTML report."""

    def export(self, data: Dict[str, Any], output_path: Optional[Path] = None) -> str:
        """Export data as HTML report."""
        filtered_data = self._filter_data(data)

        html_content = self._generate_html_report(filtered_data)

        if output_path:
            output_path.write_text(html_content, encoding="utf-8")

        return html_content

    def _generate_html_report(self, data: Dict[str, Any]) -> str:
        """Generate HTML report from data."""
        html_parts = []

        # HTML header
        html_parts.append(
            """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Python Mastery Hub - Learning Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
                .container { max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }
                h1 { color: #2E86AB; border-bottom: 2px solid #2E86AB; padding-bottom: 10px; }
                h2 { color: #27AE60; margin-top: 30px; }
                h3 { color: #8E44AD; }
                .stat-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
                .stat-box { background: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 4px solid #2E86AB; }
                .stat-box h4 { margin: 0 0 10px 0; color: #2E86AB; }
                .stat-box .value { font-size: 24px; font-weight: bold; color: #333; }
                table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                th, td { padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }
                th { background: #f8f9fa; font-weight: bold; }
                .progress-bar { background: #ddd; height: 20px; border-radius: 10px; overflow: hidden; }
                .progress-fill { background: #27AE60; height: 100%; transition: width 0.3s; }
                .achievement { background: #FFF3CD; border: 1px solid #FFE69C; padding: 10px; margin: 10px 0; border-radius: 5px; }
                .footer { text-align: center; margin-top: 40px; color: #666; font-size: 12px; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Python Mastery Hub - Learning Report</h1>
                <p><strong>Generated:</strong> {}</p>
        """.format(
                datetime.now().strftime("%B %d, %Y at %I:%M %p")
            )
        )

        # User information
        if "user_info" in data:
            html_parts.append(self._generate_user_info_html(data["user_info"]))

        # Statistics overview
        if "statistics" in data:
            html_parts.append(self._generate_statistics_html(data["statistics"]))

        # Progress data
        if "progress" in data:
            html_parts.append(self._generate_progress_html(data["progress"]))

        # Achievements
        if "achievements" in data:
            html_parts.append(self._generate_achievements_html(data["achievements"]))

        # Learning history
        if "learning_history" in data:
            html_parts.append(
                self._generate_learning_history_html(data["learning_history"])
            )

        # HTML footer
        html_parts.append(
            """
                <div class="footer">
                    <p>This report was generated by Python Mastery Hub</p>
                    <p>For questions about your learning progress, contact support@pythonmasteryhub.com</p>
                </div>
            </div>
        </body>
        </html>
        """
        )

        return "".join(html_parts)

    def _generate_user_info_html(self, user_info: Dict[str, Any]) -> str:
        """Generate user information HTML section."""
        html = "<h2>User Information</h2>"
        html += "<div class='stat-grid'>"

        for key, value in user_info.items():
            if isinstance(value, datetime):
                value = self._format_datetime(value)

            display_key = key.replace("_", " ").title()
            html += f"""
            <div class='stat-box'>
                <h4>{display_key}</h4>
                <div class='value'>{value}</div>
            </div>
            """

        html += "</div>"
        return html

    def _generate_statistics_html(self, statistics: Dict[str, Any]) -> str:
        """Generate statistics HTML section."""
        html = "<h2>Learning Statistics</h2>"
        html += "<div class='stat-grid'>"

        # Key statistics
        key_stats = {
            "completed_topics": "Topics Completed",
            "total_time_minutes": "Total Study Time (min)",
            "current_streak": "Current Streak (days)",
            "longest_streak": "Longest Streak (days)",
            "modules_started": "Modules Started",
            "average_score": "Average Score",
        }

        for key, label in key_stats.items():
            if key in statistics:
                value = statistics[key]
                if key == "average_score" and isinstance(value, (int, float)):
                    value = f"{value * 100:.1f}%"

                html += f"""
                <div class='stat-box'>
                    <h4>{label}</h4>
                    <div class='value'>{value}</div>
                </div>
                """

        html += "</div>"
        return html

    def _generate_progress_html(self, progress: Dict[str, Any]) -> str:
        """Generate progress HTML section."""
        html = "<h2>Module Progress</h2>"

        if "modules" in progress:
            html += "<table>"
            html += "<tr><th>Module</th><th>Progress</th><th>Completed</th><th>Total</th><th>Time Spent</th></tr>"

            for module_id, module_data in progress["modules"].items():
                percentage = module_data.get("percentage", 0)
                completed = module_data.get("completed", 0)
                total = module_data.get("total", 0)
                time_spent = module_data.get("time_spent", 0)

                html += f"""
                <tr>
                    <td>{module_id.replace('_', ' ').title()}</td>
                    <td>
                        <div class='progress-bar'>
                            <div class='progress-fill' style='width: {percentage}%;'></div>
                        </div>
                        {percentage:.1f}%
                    </td>
                    <td>{completed}</td>
                    <td>{total}</td>
                    <td>{time_spent} min</td>
                </tr>
                """

            html += "</table>"

        return html

    def _generate_achievements_html(self, achievements: List[Dict[str, Any]]) -> str:
        """Generate achievements HTML section."""
        html = "<h2>Achievements</h2>"

        if not achievements:
            html += "<p>No achievements earned yet.</p>"
            return html

        for achievement in achievements:
            badge = achievement.get("badge", "🏆")
            name = achievement.get("name", "Unknown Achievement")
            description = achievement.get("description", "")
            tier = achievement.get("tier", "Bronze")
            points = achievement.get("points", 0)
            earned_date = achievement.get("earned_date", "")

            if earned_date:
                earned_date = self._format_datetime(earned_date)

            html += f"""
            <div class='achievement'>
                <h4>{badge} {name} ({tier})</h4>
                <p>{description}</p>
                <p><strong>Points:</strong> {points} | <strong>Earned:</strong> {earned_date}</p>
            </div>
            """

        return html

    def _generate_learning_history_html(
        self, learning_history: List[Dict[str, Any]]
    ) -> str:
        """Generate learning history HTML section."""
        html = "<h2>Recent Learning Activity</h2>"

        if not learning_history:
            html += "<p>No learning history available.</p>"
            return html

        # Show only recent 20 entries
        recent_history = (
            learning_history[-20:] if len(learning_history) > 20 else learning_history
        )

        html += "<table>"
        html += "<tr><th>Date</th><th>Module</th><th>Topic</th><th>Action</th><th>Duration</th></tr>"

        for entry in recent_history:
            date = self._format_datetime(entry.get("date", ""))
            module = entry.get("module_id", "").replace("_", " ").title()
            topic = entry.get("topic_name", "")
            action = entry.get("action", "")
            duration = entry.get("duration", 0)

            html += f"""
            <tr>
                <td>{date}</td>
                <td>{module}</td>
                <td>{topic}</td>
                <td>{action}</td>
                <td>{duration} min</td>
            </tr>
            """

        html += "</table>"
        return html


class ExcelExporter(DataExporter):
    """Export data to Excel format (requires openpyxl)."""

    def export(self, data: Dict[str, Any], output_path: Optional[Path] = None) -> bytes:
        """Export data to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        filtered_data = self._filter_data(data)

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add sheets for different data types
        if "user_info" in filtered_data:
            self._add_user_info_sheet(wb, filtered_data["user_info"])

        if "progress" in filtered_data:
            self._add_progress_sheet(wb, filtered_data["progress"])

        if "achievements" in filtered_data:
            self._add_achievements_sheet(wb, filtered_data["achievements"])

        if "statistics" in filtered_data:
            self._add_statistics_sheet(wb, filtered_data["statistics"])

        if "learning_history" in filtered_data:
            self._add_learning_history_sheet(wb, filtered_data["learning_history"])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        result = output.getvalue()

        if output_path:
            with output_path.open("wb") as f:
                f.write(result)

        return result

    def _add_user_info_sheet(self, wb, user_info: Dict[str, Any]):
        """Add user information sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return

        ws = wb.create_sheet("User Information")

        # Headers
        ws["A1"] = "Field"
        ws["B1"] = "Value"

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )

        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill

        # Add data
        row = 2
        for key, value in user_info.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            if isinstance(value, datetime):
                ws[f"B{row}"] = self._format_datetime(value)
            else:
                ws[f"B{row}"] = str(value)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _add_progress_sheet(self, wb, progress: Dict[str, Any]):
        """Add progress sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return

        ws = wb.create_sheet("Progress")

        # Module progress section
        if "modules" in progress:
            # Headers
            headers = [
                "Module",
                "Completed Topics",
                "Total Topics",
                "Percentage",
                "Time Spent (min)",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                )

            # Data
            row = 2
            for module_id, module_data in progress["modules"].items():
                ws.cell(row=row, column=1, value=module_id.replace("_", " ").title())
                ws.cell(row=row, column=2, value=module_data.get("completed", 0))
                ws.cell(row=row, column=3, value=module_data.get("total", 0))
                ws.cell(
                    row=row, column=4, value=f"{module_data.get('percentage', 0):.1f}%"
                )
                ws.cell(row=row, column=5, value=module_data.get("time_spent", 0))
                row += 1

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _add_achievements_sheet(self, wb, achievements: List[Dict[str, Any]]):
        """Add achievements sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return

        ws = wb.create_sheet("Achievements")

        # Headers
        headers = ["Name", "Description", "Category", "Tier", "Points", "Earned Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
            )

        # Data
        row = 2
        for achievement in achievements:
            ws.cell(row=row, column=1, value=achievement.get("name", ""))
            ws.cell(row=row, column=2, value=achievement.get("description", ""))
            ws.cell(row=row, column=3, value=achievement.get("category", ""))
            ws.cell(row=row, column=4, value=achievement.get("tier", ""))
            ws.cell(row=row, column=5, value=achievement.get("points", 0))
            ws.cell(
                row=row,
                column=6,
                value=self._format_datetime(achievement.get("earned_date", "")),
            )
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 20

    def _add_statistics_sheet(self, wb, statistics: Dict[str, Any]):
        """Add statistics sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return

        ws = wb.create_sheet("Statistics")

        # Headers
        ws["A1"] = "Metric"
        ws["B1"] = "Value"

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )

        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill

        # Add data
        row = 2
        for key, value in statistics.items():
            if isinstance(value, dict):
                # Handle nested dictionaries
                for sub_key, sub_value in value.items():
                    ws[f"A{row}"] = f"{key}.{sub_key}".replace("_", " ").title()
                    ws[f"B{row}"] = str(sub_value)
                    row += 1
            elif isinstance(value, list):
                ws[f"A{row}"] = key.replace("_", " ").title()
                ws[f"B{row}"] = ", ".join(map(str, value))
                row += 1
            else:
                ws[f"A{row}"] = key.replace("_", " ").title()
                ws[f"B{row}"] = str(value)
                row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _add_learning_history_sheet(self, wb, learning_history: List[Dict[str, Any]]):
        """Add learning history sheet."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return

        ws = wb.create_sheet("Learning History")

        # Headers
        headers = ["Date", "Module", "Topic", "Action", "Duration (min)", "Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
            )

        # Data
        row = 2
        for entry in learning_history:
            ws.cell(
                row=row, column=1, value=self._format_datetime(entry.get("date", ""))
            )
            ws.cell(
                row=row,
                column=2,
                value=entry.get("module_id", "").replace("_", " ").title(),
            )
            ws.cell(row=row, column=3, value=entry.get("topic_name", ""))
            ws.cell(row=row, column=4, value=entry.get("action", ""))
            ws.cell(row=row, column=5, value=entry.get("duration", 0))
            score = entry.get("score")
            ws.cell(row=row, column=6, value=f"{score*100:.1f}%" if score else "")
            row += 1

        # Adjust column widths
        for col_letter in ["A", "B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 15


class ZipExporter:
    """Export data as a ZIP archive containing multiple formats."""

    def __init__(self, config: Optional[ExportConfig] = None):
        self.config = config or ExportConfig()

    def export(self, data: Dict[str, Any], output_path: Optional[Path] = None) -> bytes:
        """Export data as ZIP archive with multiple formats."""

        # Create exporters
        csv_exporter = CSVExporter(self.config)
        json_exporter = JSONExporter(self.config)
        html_exporter = HTMLReportExporter(self.config)

        # Create ZIP file in memory
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Add CSV export
            csv_data = csv_exporter.export(data)
            zip_file.writestr("learning_data.csv", csv_data)

            # Add JSON export
            json_data = json_exporter.export(data)
            zip_file.writestr("learning_data.json", json_data)

            # Add HTML report
            html_data = html_exporter.export(data)
            zip_file.writestr("learning_report.html", html_data)

            # Add Excel export if possible
            try:
                excel_exporter = ExcelExporter(self.config)
                excel_data = excel_exporter.export(data)
                zip_file.writestr("learning_data.xlsx", excel_data)
            except ImportError:
                logger.warning("Excel export skipped - openpyxl not available")

            # Add README
            readme_content = self._generate_readme()
            zip_file.writestr("README.txt", readme_content)

        result = zip_buffer.getvalue()

        if output_path:
            with output_path.open("wb") as f:
                f.write(result)

        return result

    def _generate_readme(self) -> str:
        """Generate README for the export archive."""
        return """
Python Mastery Hub - Learning Data Export
==========================================

This archive contains your learning data from Python Mastery Hub in multiple formats:

Files included:
- learning_data.csv: Your learning data in CSV format for spreadsheet applications
- learning_data.json: Complete data in JSON format for programmatic access
- learning_report.html: Beautiful HTML report that you can open in your web browser
- learning_data.xlsx: Excel format (if available)
- README.txt: This file

Data included:
- User information and profile data
- Learning progress across all modules
- Achievement history and badges earned
- Detailed learning statistics
- Complete learning activity history

Export date: {}
Generated by: Python Mastery Hub

For questions about your data or this export, contact support@pythonmasteryhub.com
        """.strip().format(
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )


# Factory functions
def create_exporter(
    format_type: str, config: Optional[ExportConfig] = None
) -> DataExporter:
    """
    Create exporter for specified format.

    Args:
        format_type: Export format ('csv', 'json', 'html', 'excel', 'zip')
        config: Export configuration

    Returns:
        Appropriate exporter instance
    """
    exporters = {
        "csv": CSVExporter,
        "json": JSONExporter,
        "html": HTMLReportExporter,
        "excel": ExcelExporter,
        "zip": ZipExporter,
    }

    exporter_class = exporters.get(format_type.lower())
    if not exporter_class:
        raise ValueError(f"Unsupported export format: {format_type}")

    return exporter_class(config)


def export_learning_data(
    data: Dict[str, Any],
    format_type: str = "json",
    output_path: Optional[Path] = None,
    config: Optional[ExportConfig] = None,
) -> Union[str, bytes]:
    """
    Convenience function to export learning data.

    Args:
        data: Learning data to export
        format_type: Export format
        output_path: Output file path
        config: Export configuration

    Returns:
        Exported data
    """
    exporter = create_exporter(format_type, config)
    return exporter.export(data, output_path)


def get_supported_formats() -> List[str]:
    """Get list of supported export formats."""
    return ["csv", "json", "html", "excel", "zip"]
